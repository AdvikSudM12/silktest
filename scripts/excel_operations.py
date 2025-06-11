import os
import pandas as pd
from openpyxl import load_workbook
import json
from pathlib import Path
from datetime import datetime

# DEBUG: –î–æ–±–∞–≤–ª—è–µ–º –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏–µ –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏ –æ–±—Ä–∞–±–æ—Ç–∫–∏ Excel
import sys
script_dir = Path(__file__).parent.parent
sys.path.append(str(script_dir))
from pyqt_app.logger_config import get_logger
debug_logger = get_logger("excel_operations")

def process_excel_errors(excel_file_path=None, error_file_path=None):
    """
    –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –æ—à–∏–±–∫–∏ –∏–∑ —Ñ–∞–π–ª–∞ —Å—Ä–∞–≤–Ω–µ–Ω–∏—è –∏ –ø–µ—Ä–µ–Ω–æ—Å–∏—Ç –ø—Ä–æ–±–ª–µ–º–Ω—ã–µ —Å—Ç—Ä–æ–∫–∏ –≤ –æ—Ç–¥–µ–ª—å–Ω—ã–π –ª–∏—Å—Ç
    
    Args:
        excel_file_path: –ü—É—Ç—å –∫ –æ—Å–Ω–æ–≤–Ω–æ–º—É Excel —Ñ–∞–π–ª—É
        error_file_path: –ü—É—Ç—å –∫ —Ñ–∞–π–ª—É —Å —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞–º–∏ —Å—Ä–∞–≤–Ω–µ–Ω–∏—è
    
    Returns:
        dict: –†–µ–∑—É–ª—å—Ç–∞—Ç –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Å –∫–ª—é—á–∞–º–∏ 'success', 'message', 'moved_count'
    """
    debug_logger.info("üìù –ù–∞—á–∏–Ω–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É Excel –æ—à–∏–±–æ–∫")
    debug_logger.debug(f"üìÑ –û—Å–Ω–æ–≤–Ω–æ–π Excel: {excel_file_path}")
    debug_logger.debug(f"üìä –§–∞–π–ª –æ—à–∏–±–æ–∫: {error_file_path}")
    
    # –ü–æ–ª—É—á–∞–µ–º –∞–±—Å–æ–ª—é—Ç–Ω—ã–π –ø—É—Ç—å –∫ –¥–∏—Ä–µ–∫—Ç–æ—Ä–∏–∏ —Å–∫—Ä–∏–ø—Ç–∞
    script_dir = Path(__file__).parent.parent
    debug_logger.debug(f"üìÇ –î–∏—Ä–µ–∫—Ç–æ—Ä–∏—è —Å–∫—Ä–∏–ø—Ç–∞: {script_dir}")
    
    # –ï—Å–ª–∏ –ø—É—Ç–∏ –Ω–µ –ø–µ—Ä–µ–¥–∞–Ω—ã, –∑–∞–≥—Ä—É–∂–∞–µ–º –∏–∑ paths.json
    if not excel_file_path:
        paths_file = script_dir / 'pyqt_app' / 'data' / 'paths.json'
        if os.path.exists(paths_file):
            try:
                with open(paths_file, 'r', encoding='utf-8') as f:
                    paths_data = json.load(f)
                    excel_file_path = paths_data.get('excel_file_path')
            except Exception as e:
                return {
                    'success': False,
                    'message': f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–≥—Ä—É–∑–∫–µ paths.json: {str(e)}",
                    'moved_count': 0
                }
    
    # –ï—Å–ª–∏ —Ñ–∞–π–ª —Å –æ—à–∏–±–∫–∞–º–∏ –Ω–µ —É–∫–∞–∑–∞–Ω, –∏—â–µ–º –ø–æ—Å–ª–µ–¥–Ω–∏–π –≤ –ø–∞–ø–∫–µ results
    if not error_file_path:
        results_dir = script_dir / 'results'
        if os.path.exists(results_dir):
            # –ò—â–µ–º –ø–æ—Å–ª–µ–¥–Ω–∏–π —Ñ–∞–π–ª —Å —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞–º–∏ —Å—Ä–∞–≤–Ω–µ–Ω–∏—è
            error_files = [f for f in os.listdir(results_dir) if f.startswith('file_comparison_results_') and f.endswith('.xlsx')]
            if error_files:
                error_files.sort(reverse=True)  # –°–æ—Ä—Ç–∏—Ä—É–µ–º –ø–æ —É–±—ã–≤–∞–Ω–∏—é (–ø–æ—Å–ª–µ–¥–Ω–∏–π —Ñ–∞–π–ª –ø–µ—Ä–≤—ã–π)
                error_file_path = results_dir / error_files[0]
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ —Ñ–∞–π–ª–æ–≤
    if not excel_file_path or not os.path.exists(excel_file_path):
        return {
            'success': False,
            'message': f"–û—Å–Ω–æ–≤–Ω–æ–π Excel —Ñ–∞–π–ª –Ω–µ –Ω–∞–π–¥–µ–Ω: {excel_file_path}",
            'moved_count': 0
        }
    
    if not error_file_path or not os.path.exists(error_file_path):        return {
            'success': False,
            'message': f"–§–∞–π–ª —Å –æ—à–∏–±–∫–∞–º–∏ –Ω–µ –Ω–∞–π–¥–µ–Ω: {error_file_path}",
            'moved_count': 0
        }

    try:
        # –ß–∏—Ç–∞–µ–º —Ñ–∞–π–ª —Å –æ—à–∏–±–∫–∞–º–∏ —Å –ø–æ–º–æ—â—å—é pandas
        error_df = pd.read_excel(error_file_path, sheet_name='–†–µ–∑—É–ª—å—Ç–∞—Ç—ã')
        
        # –°–æ–±–∏—Ä–∞–µ–º —Å–ø–∏—Å–æ–∫ –Ω–∞–∑–≤–∞–Ω–∏–π —Ç—Ä–µ–∫–æ–≤ —Å –æ—à–∏–±–∫–∞–º–∏
        error_tracks = set()
        for _, row in error_df.iterrows():
            if pd.notna(row['–ù–∞–∑–≤–∞–Ω–∏–µ –≤ Excel']):  # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –∑–Ω–∞—á–µ–Ω–∏–µ –Ω–µ NaN
                track_name = str(row['–ù–∞–∑–≤–∞–Ω–∏–µ –≤ Excel']).strip()
                error_tracks.add(track_name)
        
        if len(error_tracks) == 0:
            return {
                'success': True,
                'message': "–ù–µ –Ω–∞–π–¥–µ–Ω–æ —Ç—Ä–µ–∫–æ–≤ —Å –æ—à–∏–±–∫–∞–º–∏ –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏",
                'moved_count': 0
            }
        
        # –û—Ç–∫—Ä—ã–≤–∞–µ–º –æ—Å–Ω–æ–≤–Ω–æ–π —Ñ–∞–π–ª
        main_wb = load_workbook(excel_file_path)
        if "–õ–∏—Å—Ç1" not in main_wb.sheetnames:
            return {
                'success': False,
                'message': "–õ–∏—Å—Ç '–õ–∏—Å—Ç1' –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ –æ—Å–Ω–æ–≤–Ω–æ–º —Ñ–∞–π–ª–µ",
                'moved_count': 0
            }
        
        main_sheet = main_wb["–õ–∏—Å—Ç1"]
        
        # –°–æ–∑–¥–∞–µ–º –∏–º—è –ª–∏—Å—Ç–∞ —Å —Ç–µ–∫—É—â–µ–π –¥–∞—Ç–æ–π
        current_date = datetime.now().strftime("%d.%m.%y")
        error_sheet_name = f"–û—à–∏–±–∫–∏_–∑–∞–≥—Ä—É–∑–∫–∏_{current_date}"
        
        # –°–æ–∑–¥–∞–µ–º –∏–ª–∏ –æ—á–∏—â–∞–µ–º –ª–∏—Å—Ç –¥–ª—è –æ—à–∏–±–æ–∫
        if error_sheet_name in main_wb.sheetnames:
            main_wb.remove(main_wb[error_sheet_name])
        error_sheet_new = main_wb.create_sheet(error_sheet_name)
          # –ö–æ–ø–∏—Ä—É–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏
        headers = [cell.value for cell in main_sheet[1]]
        error_sheet_new.append(headers)
        
        # –ü–µ—Ä–µ–±–∏—Ä–∞–µ–º —Å—Ç—Ä–æ–∫–∏ –∏ –ø–µ—Ä–µ–Ω–æ—Å–∏–º —Ç–æ–ª—å–∫–æ —Ç–µ, —á—Ç–æ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ –æ—à–∏–±–æ–∫
        rows_to_delete = []
        
        for row_idx, row in enumerate(main_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0]:
                current_track = str(row[0]).strip()
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —Ç–µ–∫—É—â–∏–π —Ç—Ä–µ–∫ –≤ —Å–ø–∏—Å–∫–µ –æ—à–∏–±–æ–∫
                if current_track in error_tracks:
                    error_sheet_new.append(row)
                    rows_to_delete.append(row_idx)
        
        # –£–¥–∞–ª—è–µ–º —Å—Ç—Ä–æ–∫–∏ —Å –∫–æ–Ω—Ü–∞, —á—Ç–æ–±—ã –Ω–µ —Å–±–∏—Ç—å –Ω—É–º–µ—Ä–∞—Ü–∏—é
        for row_idx in reversed(rows_to_delete):
            main_sheet.delete_rows(row_idx)
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –∏–∑–º–µ–Ω–µ–Ω–∏—è
        main_wb.save(excel_file_path)
        
        return {
            'success': True,
            'message': f"–û–ø–µ—Ä–∞—Ü–∏—è —É—Å–ø–µ—à–Ω–æ –≤—ã–ø–æ–ª–Ω–µ–Ω–∞! –ü–µ—Ä–µ–Ω–µ—Å–µ–Ω–æ {len(rows_to_delete)} —Å—Ç—Ä–æ–∫ —Å –æ—à–∏–±–∫–∞–º–∏ –≤ –ª–∏—Å—Ç '{error_sheet_name}'",
            'moved_count': len(rows_to_delete),
            'error_sheet_name': error_sheet_name
        }

    except Exception as e:
        return {
            'success': False,
            'message': f"–ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞: {str(e)}",
            'moved_count': 0
        }


def process_excel_errors_interactive():
    """–ò–Ω—Ç–µ—Ä–∞–∫—Ç–∏–≤–Ω–∞—è –≤–µ—Ä—Å–∏—è –¥–ª—è –∑–∞–ø—É—Å–∫–∞ –∏–∑ –∫–æ–º–∞–Ω–¥–Ω–æ–π —Å—Ç—Ä–æ–∫–∏"""
    main_file = input("–í–≤–µ–¥–∏—Ç–µ –ø—É—Ç—å –∫ –æ—Å–Ω–æ–≤–Ω–æ–º—É Excel —Ñ–∞–π–ª—É: ")
    error_file = input("–í–≤–µ–¥–∏—Ç–µ –ø—É—Ç—å –∫ —Ñ–∞–π–ª—É —Å –æ—à–∏–±–∫–∞–º–∏ (–∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ Enter –¥–ª—è –∞–≤—Ç–æ–ø–æ–∏—Å–∫–∞): ")
    
    if not error_file.strip():
        error_file = None
    
    result = process_excel_errors(main_file, error_file)
    
    if result['success']:
        print(f"\n{result['message']}")
    else:
        print(f"–û—à–∏–±–∫–∞: {result['message']}")


if __name__ == "__main__":
    process_excel_errors_interactive()
