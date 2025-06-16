import pandas as pd
import os
from difflib import SequenceMatcher
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import unicodedata
import re

# DEBUG: Добавляем логирование для отладки процесса сравнения файлов
# Импортируем логгер из pyqt_app
import sys
script_dir = Path(__file__).parent.parent
sys.path.append(str(script_dir))
from pyqt_app.logger_config import get_logger
debug_logger = get_logger("compare_files")

def normalize_filename(filename):
    """Нормализует имя файла для корректного сравнения"""
    if not filename or pd.isna(filename) or str(filename).strip() == '':
        return ''
    
    # Преобразуем в строку и убираем пробелы по краям
    filename = str(filename).strip()
    
    # Нормализация Unicode
    filename = unicodedata.normalize('NFKC', filename)
    
    # Обработка пробелов вокруг специальных символов
    # Убираем пробелы перед скобками, точками и другими спецсимволами
    filename = re.sub(r'\s*([\(\)\[\]\{\}\.,\-_])\s*', r'\1', filename)
    
    # Заменяем множественные пробелы на один
    filename = re.sub(r'\s+', ' ', filename)
    
    # Приводим к нижнему регистру
    filename = filename.lower()
    
    # Отдельно обрабатываем расширение файла
    name, ext = os.path.splitext(filename)
    if ext:
        # Убираем точку из расширения для сравнения
        ext = ext[1:] if ext.startswith('.') else ext
        # Убираем пробелы вокруг точки расширения
        return f"{name.strip()}.{ext.strip()}"
    
    return filename.strip()

def calculate_similarity(a, b):
    """Вычисляет процент сходства между двумя строками"""
    if not a or not b:
        return 0
    
    # Нормализуем строки перед сравнением
    a = normalize_filename(a)
    b = normalize_filename(b)
    
    # Разделяем имя и расширение
    a_name, a_ext = os.path.splitext(a)
    b_name, b_ext = os.path.splitext(b)
    
    # Если расширения разные, уменьшаем сходство
    if a_ext.lower() != b_ext.lower():
        return 0
    
    # Сравниваем только имена файлов без расширений
    similarity = SequenceMatcher(None, a_name, b_name).ratio() * 100
    
    return similarity

def find_closest_match(filename, file_list):
    """Находит самое похожее название файла и процент сходства"""
    if not filename or pd.isna(filename) or str(filename).strip() == '':
        return '', 0
    
    normalized_filename = normalize_filename(filename)
    if not normalized_filename:
        return '', 0
    
    # Выводим отладочную информацию
    debug_logger.info(f"\n🔍 Искомый файл: '{filename}'")
    debug_logger.info(f"🔧 Нормализованный файл: '{normalized_filename}'")
    
    # Разделяем имя и расширение искомого файла
    filename_name, filename_ext = os.path.splitext(normalized_filename)
    
    max_similarity = 0
    closest_match = None
    closest_match_original = None
    
    for original_file in file_list:
        normalized_file = normalize_filename(original_file)
        # Проверяем расширение
        _, file_ext = os.path.splitext(normalized_file)
        
        # Если расширения разные, пропускаем файл
        if filename_ext.lower() != file_ext.lower():
            continue
            
        similarity = calculate_similarity(normalized_filename, normalized_file)
        if similarity > max_similarity:
            max_similarity = similarity
            closest_match = normalized_file
            closest_match_original = original_file
            
            # Выводим информацию о найденном совпадении
            debug_logger.debug(f"\n🎯 Найдено лучшее совпадение:")
            debug_logger.debug(f"📝 Искомый файл     : '{filename}'")
            debug_logger.debug(f"🔧 Нормализованный  : '{normalized_filename}'")
            debug_logger.debug(f"📁 Найденный файл   : '{original_file}'")
            debug_logger.debug(f"🔧 Нормализованный  : '{normalized_file}'")
            debug_logger.debug(f"📊 Процент сходства : {similarity}%")
    
    return closest_match_original, max_similarity

def find_char_differences(str1, str2):
    """Находит различия между двумя строками"""
    # Нормализуем строки перед сравнением
    str1 = normalize_filename(str1)
    str2 = normalize_filename(str2)
    
    char_differences = []
    
    # Сравниваем символы
    for i, (c1, c2) in enumerate(zip(str1, str2)):
        if c1 != c2:
            # Показываем оригинальные символы для лучшего понимания различий
            char_differences.append(f"Позиция {i+1}: '{c1}' vs '{c2}'")
    
    # Проверяем разницу в длине
    if len(str1) != len(str2):
        min_len = min(len(str1), len(str2))
        # Показываем лишние символы
        if len(str1) > min_len:
            extra = str1[min_len:]
            char_differences.append(f"Лишние символы в первой строке: '{extra}'")
        if len(str2) > min_len:
            extra = str2[min_len:]
            char_differences.append(f"Лишние символы во второй строке: '{extra}'")
    
    return '; '.join(char_differences) if char_differences else ''

def compare_files_with_excel(excel_file_path=None, directory_path=None):
    """
    Сравнивает файлы из Excel с реальными файлами в директории
    
    Args:
        excel_file_path: Путь к Excel файлу (если не указан, берется из paths.json)
        directory_path: Путь к директории с файлами (если не указан, берется из paths.json)
    
    Returns:
        dict: Результат сравнения с ключами 'success', 'results_file', 'error_count', 'message'
    """
    debug_logger.info("🔍 Начинаем сравнение файлов с Excel")
    debug_logger.debug(f"📄 Excel файл: {excel_file_path}")
    debug_logger.debug(f"📁 Директория: {directory_path}")
    
    import json
    
    # Получаем абсолютный путь к директории скрипта
    script_dir = Path(__file__).parent.parent
    debug_logger.debug(f"📂 Директория скрипта: {script_dir}")
    
    # Если пути не переданы, загружаем из paths.json
    if not excel_file_path or not directory_path:
        debug_logger.info("🔄 Загружаем пути из paths.json")
        paths_file = script_dir / 'pyqt_app' / 'data' / 'paths.json'
        debug_logger.debug(f"📍 Путь к paths.json: {paths_file}")
        
        if os.path.exists(paths_file):
            debug_logger.success("✅ Файл paths.json найден")
            try:
                with open(paths_file, 'r', encoding='utf-8') as f:
                    paths_data = json.load(f)
                    debug_logger.debug(f"📊 Данные из paths.json: {paths_data}")
                    
                    if not excel_file_path:
                        excel_file_path = paths_data.get('excel_file_path')
                        debug_logger.debug(f"📄 Загружен Excel путь: {excel_file_path}")
                    if not directory_path:
                        directory_path = paths_data.get('directory_path')
                        debug_logger.debug(f"📁 Загружен путь директории: {directory_path}")
            except Exception as e:
                debug_logger.error(f"❌ Ошибка при загрузке paths.json: {str(e)}")
                return {
                    'success': False,
                    'message': f"Ошибка при загрузке paths.json: {str(e)}",
                    'results_file': None,
                    'error_count': 0
                }
        else:
            debug_logger.warning("⚠️ Файл paths.json не найден")
    
    # Проверяем наличие путей
    debug_logger.info("🔍 Проверяем наличие путей")
    if not excel_file_path:
        debug_logger.error("❌ Не указан путь к Excel файлу")
        return {
            'success': False,
            'message': "Не указан путь к Excel файлу",
            'results_file': None,
            'error_count': 0
        }
    
    if not directory_path:
        debug_logger.error("❌ Не указан путь к директории с файлами")
        return {
            'success': False,
            'message': "Не указан путь к директории с файлами",
            'results_file': None,
            'error_count': 0
        }
    
    debug_logger.info("🔍 Проверяем существование файлов")
    if not os.path.exists(excel_file_path):
        debug_logger.error(f"❌ Excel файл не найден: {excel_file_path}")
        return {
            'success': False,
            'message': f"Excel файл не найден: {excel_file_path}",
            'results_file': None,
            'error_count': 0
        }
        
    if not os.path.exists(directory_path):
        debug_logger.error(f"❌ Директория не найдена: {directory_path}")
        return {
            'success': False,
            'message': f"Директория не найдена: {directory_path}",
            'results_file': None,
            'error_count': 0
        }    
    
    debug_logger.success("✅ Все пути проверены и существуют")
    
    # Создаем директорию results, если её нет
    results_dir = script_dir / 'results'
    debug_logger.debug(f"📁 Директория результатов: {results_dir}")
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)
        debug_logger.info("📁 Создана директория для результатов")

    # Получаем список реальных файлов в директории
    debug_logger.info("📋 Получаем список файлов в директории")
    actual_files = [f for f in os.listdir(directory_path)]
    debug_logger.info(f"📊 Найдено {len(actual_files)} файлов в директории")
    debug_logger.debug(f"📝 Список файлов: {actual_files[:10]}...")  # Показываем только первые 10
    
    try:
        debug_logger.info("📖 Читаем Excel файл")
        # Читаем только Лист1 из Excel файла
        df = pd.read_excel(excel_file_path, sheet_name='Лист1', engine='openpyxl')
        debug_logger.success(f"✅ Excel файл прочитан, строк: {len(df)}")
    except Exception as e:        
        debug_logger.error(f"❌ Ошибка при чтении Excel файла: {str(e)}")
        return {
            'success': False,
            'message': f"Ошибка при чтении Excel файла: {str(e)}",
            'results_file': None,
            'error_count': 0
        }
    
    # Проверяем наличие необходимых столбцов
    required_columns = ['track (titel)', 'cover (titel)']
    if not all(col in df.columns for col in required_columns):
        return {
            'success': False,
            'message': f"В Excel файле отсутствуют необходимые столбцы: {required_columns}",
            'results_file': None,
            'error_count': 0
        }
    
    # Логируем информацию о колонках Excel файла
    debug_logger.info(f"📋 Найденные колонки в Excel файле: {list(df.columns)}")
    if 'release_name' in df.columns:
        debug_logger.info("✅ Колонка 'release_name' найдена в Excel файле")
        # Проверяем сколько строк имеют заполненную колонку release_name
        filled_releases = df['release_name'].notna().sum()
        total_rows = len(df)
        debug_logger.info(f"📊 Заполненных релизов: {filled_releases}/{total_rows}")
        if filled_releases > 0:
            unique_releases = df['release_name'].dropna().unique()
            debug_logger.info(f"🎵 Уникальные релизы: {list(unique_releases)}")
    else:
        debug_logger.warning("⚠️ Колонка 'release_name' НЕ найдена в Excel файле - все файлы будут отнесены к 'Без указания релиза'")

    # Создаем списки для хранения результатов
    all_results = []
    errors_only = []
    statistics = {
        'total_excel_tracks': 0,
        'total_excel_covers': 0,
        'total_actual_files': len(actual_files),
        'perfect_matches': 0,
        'partial_matches': 0,
        'no_matches': 0,
        'tracks_processed': 0,
        'covers_processed': 0,
        'similarity_ranges': {
            '90-100%': 0,
            '80-89%': 0,
            '50-79%': 0,
            '0-49%': 0
        }
    }

    # Создаем множества для отслеживания использованных файлов
    used_files = set()
    excel_files = set()

    # Обрабатываем каждую строку в Excel
    debug_logger.info("🔄 Обрабатываем файлы из Excel")
    total_rows = len(df)
    debug_logger.info(f"📊 Всего строк для обработки: {total_rows}")
    
    for index, row in df.iterrows():
        # Логируем прогресс каждые 10 строк
        if (index + 1) % 10 == 0 or index == 0:
            debug_logger.debug(f"📈 Обработано строк: {index + 1}/{total_rows}")
        # Проверяем треки
        track_name = row['track (titel)']
        if pd.notna(track_name) and str(track_name).strip():
            statistics['total_excel_tracks'] += 1
            statistics['tracks_processed'] += 1
            normalized_track = normalize_filename(str(track_name))
            excel_files.add(normalized_track)
            
            closest_track, track_similarity = find_closest_match(track_name, actual_files)
            
            # Категоризируем по проценту сходства
            if track_similarity >= 90:
                statistics['similarity_ranges']['90-100%'] += 1
                if track_similarity == 100:
                    statistics['perfect_matches'] += 1
                else:
                    statistics['partial_matches'] += 1
            elif track_similarity >= 80:
                statistics['similarity_ranges']['80-89%'] += 1
                statistics['partial_matches'] += 1
            elif track_similarity >= 50:
                statistics['similarity_ranges']['50-79%'] += 1
                statistics['partial_matches'] += 1
            else:
                statistics['similarity_ranges']['0-49%'] += 1
                statistics['no_matches'] += 1
            
            if closest_track:
                used_files.add(normalize_filename(closest_track))
            
            differences = find_char_differences(str(track_name), str(closest_track)) if closest_track else 'Файл не найден'
            
            result_entry = {
                    'Тип файла': 'Трек',
                    'Название в Excel': track_name,
                    'Найден в папке': closest_track if closest_track else 'Не найден',
                    'Ближайшее совпадение': closest_track if closest_track else '',
                'Процент сходства': round(track_similarity, 2) if track_similarity > 0 else 0,
                'Различия': differences,
                'Статус': 'Точное соответствие' if track_similarity == 100 else 
                         'Частичное соответствие' if track_similarity >= 50 else 
                         'Не найден'
            }
            
            all_results.append(result_entry)
            
            # Добавляем в errors_only только если есть проблемы
            if track_similarity < 100:
                errors_only.append(result_entry)
                # Логируем найденную ошибку
                if track_similarity == 0:
                    debug_logger.warning(f"❌ Трек НЕ НАЙДЕН: '{track_name}'")
                elif track_similarity < 50:
                    debug_logger.warning(f"🔴 Трек низкое сходство ({track_similarity}%): '{track_name}' → '{closest_track}'")
                elif track_similarity < 90:
                    debug_logger.debug(f"🟡 Трек среднее сходство ({track_similarity}%): '{track_name}' → '{closest_track}'")

        # Проверяем обложки
        cover_name = row['cover (titel)']
        if pd.notna(cover_name) and str(cover_name).strip():
            statistics['total_excel_covers'] += 1
            statistics['covers_processed'] += 1
            normalized_cover = normalize_filename(str(cover_name))
            excel_files.add(normalized_cover)
            
            closest_cover, cover_similarity = find_closest_match(cover_name, actual_files)
            
            # Категоризируем по проценту сходства
            if cover_similarity >= 90:
                statistics['similarity_ranges']['90-100%'] += 1
                if cover_similarity == 100:
                    statistics['perfect_matches'] += 1
                else:
                    statistics['partial_matches'] += 1
            elif cover_similarity >= 80:
                statistics['similarity_ranges']['80-89%'] += 1
                statistics['partial_matches'] += 1
            elif cover_similarity >= 50:
                statistics['similarity_ranges']['50-79%'] += 1
                statistics['partial_matches'] += 1
            else:
                statistics['similarity_ranges']['0-49%'] += 1
                statistics['no_matches'] += 1
            
            if closest_cover:
                used_files.add(normalize_filename(closest_cover))
            
            differences = find_char_differences(str(cover_name), str(closest_cover)) if closest_cover else 'Файл не найден'
            
            result_entry = {
                    'Тип файла': 'Обложка',
                    'Название в Excel': cover_name,
                    'Найден в папке': closest_cover if closest_cover else 'Не найден',
                    'Ближайшее совпадение': closest_cover if closest_cover else '',
                'Процент сходства': round(cover_similarity, 2) if cover_similarity > 0 else 0,
                'Различия': differences,
                'Статус': 'Точное соответствие' if cover_similarity == 100 else 
                         'Частичное соответствие' if cover_similarity >= 50 else 
                         'Не найден'
            }
            
            all_results.append(result_entry)

            # Добавляем в errors_only только если есть проблемы
            if cover_similarity < 100:
                errors_only.append(result_entry)
                # Логируем найденную ошибку
                if cover_similarity == 0:
                    debug_logger.warning(f"❌ Обложка НЕ НАЙДЕНА: '{cover_name}'")
                elif cover_similarity < 50:
                    debug_logger.warning(f"🔴 Обложка низкое сходство ({cover_similarity}%): '{cover_name}' → '{closest_cover}'")
                elif cover_similarity < 90:
                    debug_logger.debug(f"🟡 Обложка среднее сходство ({cover_similarity}%): '{cover_name}' → '{closest_cover}'")

    # Находим неиспользованные файлы
    debug_logger.info("🔍 Ищем неиспользованные файлы в директории")
    unused_files = []
    for file in actual_files:
        normalized_file = normalize_filename(file)
        if normalized_file not in used_files:
            unused_files.append({'Файл в папке': file, 'Статус': 'Не найден в Excel'})
            debug_logger.debug(f"📁 Неиспользованный файл: '{file}'")
    
    debug_logger.info(f"📊 Итоговая статистика обработки:")
    debug_logger.info(f"   📄 Всего файлов в Excel: {statistics['total_excel_tracks'] + statistics['total_excel_covers']}")
    debug_logger.info(f"   🎯 Точных совпадений: {statistics['perfect_matches']}")
    debug_logger.info(f"   🟡 Частичных совпадений: {statistics['partial_matches']}")
    debug_logger.info(f"   ❌ Не найдено: {statistics['no_matches']}")
    debug_logger.info(f"   📁 Неиспользованных файлов: {len(unused_files)}")

    # Создаем DataFrames
    all_results_df = pd.DataFrame(all_results)
    
    # Для листа "Только ошибки" - если ошибок нет, добавляем сообщение
    if len(errors_only) == 0:
        errors_only_df = pd.DataFrame([['Все файлы из Excel найдены в директории', '', '', '', '', '', '']], 
                                    columns=['Тип файла', 'Название в Excel', 'Найден в папке', 'Ближайшее совпадение', 'Процент сходства', 'Различия', 'Статус'])
    else:
        errors_only_df = pd.DataFrame(errors_only)
    
    unused_files_df = pd.DataFrame(unused_files)
    
    # Создаем сводную статистику
    total_files_excel = statistics['total_excel_tracks'] + statistics['total_excel_covers']
    success_rate = (statistics['perfect_matches'] / total_files_excel * 100) if total_files_excel > 0 else 0
    
    # Создаем статистику по релизам
    release_stats = {}
    for index, row in df.iterrows():
        release_name = row.get('release_name', 'Без указания релиза')
        if pd.isna(release_name) or str(release_name).strip() == '':
            release_name = 'Без указания релиза'
        
        if release_name not in release_stats:
            release_stats[release_name] = {
                'total_files': 0,
                'found_files': 0,
                'missing_files': 0
            }
        
        # Подсчитываем треки для релиза
        track_name = row['track (titel)']
        if pd.notna(track_name) and str(track_name).strip():
            release_stats[release_name]['total_files'] += 1
            closest_track, track_similarity = find_closest_match(track_name, actual_files)
            if track_similarity >= 50:  # Считаем найденным если сходство >= 50%
                release_stats[release_name]['found_files'] += 1
            else:
                release_stats[release_name]['missing_files'] += 1
        
        # Подсчитываем обложки для релиза
        cover_name = row['cover (titel)']
        if pd.notna(cover_name) and str(cover_name).strip():
            release_stats[release_name]['total_files'] += 1  
            closest_cover, cover_similarity = find_closest_match(cover_name, actual_files)
            if cover_similarity >= 50:  # Считаем найденным если сходство >= 50%
                release_stats[release_name]['found_files'] += 1
            else:
                release_stats[release_name]['missing_files'] += 1
    
    # Формируем краткую сводку согласно образцу
    executive_summary = [
        ['Сводка по проверке файлов', ''],
        ['', ''],
        ['Общая статистика', ''],
        ['Всего файлов в Excel:', total_files_excel],
        ['Найдено совпадений:', statistics['perfect_matches'] + statistics['partial_matches']],
        ['Отсутствующие файлы:', statistics['no_matches']],
        ['Неиспользуемые файлы в директории:', len(unused_files)],
        ['', ''],
        ['Статистика по типам файлов', ''],
        ['Найдено треков:', sum(1 for result in all_results if result['Тип файла'] == 'Трек' and result['Процент сходства'] >= 50)],
        ['Найдено обложек:', sum(1 for result in all_results if result['Тип файла'] == 'Обложка' and result['Процент сходства'] >= 50)],
        ['Отсутствующие треки:', sum(1 for result in all_results if result['Тип файла'] == 'Трек' and result['Процент сходства'] < 50)],
        ['Отсутствующие обложки:', sum(1 for result in all_results if result['Тип файла'] == 'Обложка' and result['Процент сходства'] < 50)],
        ['', ''],
        ['Статистика по релизам', ''],
        ['Релиз', 'Всего файлов', 'Найдено', 'Отсутствует', 'Процент найденных']
    ]
    
    # Логируем статистику по релизам перед добавлением в отчет
    debug_logger.info("🎵 Формируем статистику по релизам:")
    for release_name, stats in sorted(release_stats.items()):
        percentage = (stats['found_files'] / stats['total_files'] * 100) if stats['total_files'] > 0 else 0
        debug_logger.info(f"   📀 '{release_name}': {stats['total_files']} файлов, найдено {stats['found_files']} ({percentage:.0f}%)")
    
    # Добавляем статистику по каждому релизу
    for release_name, stats in sorted(release_stats.items()):
        percentage = (stats['found_files'] / stats['total_files'] * 100) if stats['total_files'] > 0 else 0
        executive_summary.append([
            release_name,
            stats['total_files'],
            stats['found_files'], 
            stats['missing_files'],
            f"{percentage:.0f}%"
        ])
    
    executive_summary_df = pd.DataFrame(executive_summary)
    
    # Читаем JWT токен менеджера из config.json
    manager_token = 'Не найден'
    try:
        config_file = script_dir / 'pyqt_app' / 'data' / 'config.json'
        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
                manager_token = config_data.get('jwt', 'Не найден')
                debug_logger.debug(f"🔑 Загружен JWT токен менеджера: {manager_token[:50]}..." if len(str(manager_token)) > 50 else f"🔑 Загружен JWT токен менеджера: {manager_token}")
    except Exception as e:
        debug_logger.warning(f"⚠️ Ошибка при чтении JWT токена менеджера: {str(e)}")
    
    # Создаем детальную статистику
    detailed_stats = [
        ['Категория', 'Количество', 'Процент'],
        ['Треки в Excel', statistics['tracks_processed'], 
         f"{(statistics['tracks_processed']/total_files_excel*100):.1f}%" if total_files_excel > 0 else "0%"],
        ['Обложки в Excel', statistics['covers_processed'], 
         f"{(statistics['covers_processed']/total_files_excel*100):.1f}%" if total_files_excel > 0 else "0%"],
        ['', '', ''],
        ['Точные совпадения (100%)', statistics['perfect_matches'], 
         f"{(statistics['perfect_matches']/total_files_excel*100):.1f}%" if total_files_excel > 0 else "0%"],
        ['Высокое сходство (90-99%)', statistics['similarity_ranges']['90-100%'] - statistics['perfect_matches'], 
         f"{((statistics['similarity_ranges']['90-100%'] - statistics['perfect_matches'])/total_files_excel*100):.1f}%" if total_files_excel > 0 else "0%"],
        ['Среднее сходство (80-89%)', statistics['similarity_ranges']['80-89%'], 
         f"{(statistics['similarity_ranges']['80-89%']/total_files_excel*100):.1f}%" if total_files_excel > 0 else "0%"],
        ['Низкое сходство (50-79%)', statistics['similarity_ranges']['50-79%'], 
         f"{(statistics['similarity_ranges']['50-79%']/total_files_excel*100):.1f}%" if total_files_excel > 0 else "0%"],
        ['Очень низкое сходство (0-49%)', statistics['similarity_ranges']['0-49%'], 
         f"{(statistics['similarity_ranges']['0-49%']/total_files_excel*100):.1f}%" if total_files_excel > 0 else "0%"],
        ['', '', ''],
        ['Неиспользованные файлы', len(unused_files), 
         f"{(len(unused_files)/statistics['total_actual_files']*100):.1f}%" if statistics['total_actual_files'] > 0 else "0%"],
        ['', '', ''],
        ['JWT токен менеджера', manager_token, '']
    ]
    
    detailed_stats_df = pd.DataFrame(detailed_stats)
    
    # Создаем рекомендации
    recommendations = []
    if statistics['perfect_matches'] == total_files_excel:
        recommendations.append(['✅ Отлично!', 'Все файлы найдены с точным соответствием'])
    else:
        if statistics['no_matches'] > 0:
            recommendations.append(['🚨 Критично', f'{statistics["no_matches"]} файлов не найдено - проверьте их наличие'])
        if statistics['similarity_ranges']['0-49%'] > 0:
            recommendations.append(['⚠️ Внимание', f'{statistics["similarity_ranges"]["0-49%"]} файлов с очень низким сходством - возможны ошибки в названиях'])
        if statistics['similarity_ranges']['50-79%'] > 0:
            recommendations.append(['📝 Рекомендация', f'{statistics["similarity_ranges"]["50-79%"]} файлов требуют проверки названий'])
        if len(unused_files) > 0:
            recommendations.append(['📁 Информация', f'{len(unused_files)} файлов в папке не указаны в Excel'])
        if statistics['partial_matches'] > statistics['perfect_matches']:
            recommendations.append(['🔧 Улучшение', 'Рекомендуется стандартизировать именование файлов'])
    
    if not recommendations:
        recommendations.append(['✅ Все хорошо', 'Проблем не обнаружено'])
    
    recommendations_df = pd.DataFrame(recommendations, columns=['Приоритет', 'Рекомендация'])
    
    # Создаем имя выходного файла с текущей датой и временем
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = results_dir / f"file_comparison_results_{timestamp}.xlsx"
    
    # Создаем папку для постоянного хранения отчетов, если её нет
    reports_archive_dir = script_dir / 'verification reports'
    debug_logger.debug(f"📁 Папка архива отчетов: {reports_archive_dir}")
    if not os.path.exists(reports_archive_dir):
        os.makedirs(reports_archive_dir)
        debug_logger.info("📁 Создана папка для архива отчетов")
    
    # Путь для дубликата отчета в архиве
    archive_output_file = reports_archive_dir / f"file_comparison_results_{timestamp}.xlsx"
    
    debug_logger.info("📊 Создаем расширенный отчет Excel с 5 листами")
    
    # Функция для сохранения отчета (будем использовать дважды)
    def save_report_to_file(file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Лист 1: Все файлы (теперь первый)
            all_results_df.to_excel(writer, index=False, sheet_name='Все файлы')
        
            # Лист 2: Только ошибки (теперь второй)
            errors_only_df.to_excel(writer, index=False, sheet_name='Только ошибки')
            
            # Лист 3: Краткая сводка
            executive_summary_df.to_excel(writer, index=False, header=False, sheet_name='Краткая сводка')
            
            # Лист 4: Статистика
            detailed_stats_df.to_excel(writer, index=False, header=False, sheet_name='Детальная статистика')
            
            # Лист 5: Рекомендации
            recommendations_df.to_excel(writer, index=False, sheet_name='Рекомендации')
            
            # Лист 6: Неиспользованные файлы (если есть)
            if len(unused_files) > 0:
                unused_files_df.to_excel(writer, index=False, sheet_name='Неиспользованные файлы')
            
            # Форматирование листов
            _format_excel_sheets(writer, all_results_df, errors_only_df, len(unused_files))
    
    # Сохраняем отчет в основную папку results (для интерфейса)
    debug_logger.info("💾 Сохраняем отчет в папку results")
    save_report_to_file(output_file)
    
    # Сохраняем дубликат в архивную папку (всегда, независимо от диалога сохранения)
    debug_logger.info("📂 Сохраняем дубликат отчета в архивную папку")
    save_report_to_file(archive_output_file)
    
    error_count = len(errors_only)
    success_message = f"Найдено {error_count} файлов с различиями" if error_count > 0 else "Все файлы соответствуют записям в Excel"
    
    result = {
        'success': True,
        'message': success_message,
        'results_file': str(output_file),
        'error_count': error_count,
        'results_data': all_results
    }

    if not result['success']:
        debug_logger.error(f"\n❌ {result['message']}")
    else:
        debug_logger.success(f"\n✅ {result['message']}")
        debug_logger.info(f"📊 Расширенный отчет сохранен в файл: {result['results_file']}")
        debug_logger.info(f"📂 Дубликат сохранен в архив: {archive_output_file}")
        debug_logger.info(f"📋 Создано листов: {'6' if len(unused_files) > 0 else '5'}")
        
        if result.get('error_count', 0) > 0:
            debug_logger.warning(f"⚠️ Найдено ошибок: {result['error_count']}")
            
            # Выводим детальную информацию о файлах с ошибками
            debug_logger.info("📝 Детали файлов с ошибками:")
            for i, error in enumerate(errors_only[:10], 1):  # Показываем первые 10 ошибок
                file_name = error.get('Название в Excel', 'Неизвестно')
                file_type = error.get('Тип файла', 'Неизвестно')
                similarity = error.get('Процент сходства', 0)
                found_file = error.get('Найден в папке', 'Не найден')
                
                if similarity == 0:
                    debug_logger.error(f"   {i}. ❌ {file_type}: '{file_name}' - НЕ НАЙДЕН")
                elif similarity < 50:
                    debug_logger.warning(f"   {i}. 🔴 {file_type}: '{file_name}' - {similarity}% сходства с '{found_file}'")
                elif similarity < 80:
                    debug_logger.info(f"   {i}. 🟡 {file_type}: '{file_name}' - {similarity}% сходства с '{found_file}'")
                else:
                    debug_logger.info(f"   {i}. 🟠 {file_type}: '{file_name}' - {similarity}% сходства с '{found_file}'")
            
            if len(errors_only) > 10:
                debug_logger.info(f"   ... и еще {len(errors_only) - 10} файлов с ошибками")
                
        else:
            debug_logger.success(f"🎉 Ошибок не найдено!")

    return result


def _format_excel_sheets(writer, all_results_df, errors_only_df, unused_files_count):
    """Форматирует листы Excel для лучшего визуального восприятия"""
    
    # Определяем стили
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    bold_font = Font(bold=True)
    
    # Форматируем лист "Краткая сводка"
    if 'Краткая сводка' in writer.sheets:
        ws = writer.sheets['Краткая сводка']
        
        # Определяем специальные строки для форматирования
        section_headers = ['Сводка по проверке файлов', 'Общая статистика', 'Статистика по типам файлов', 'Статистика по релизам']
        table_header = ['Релиз', 'Всего файлов', 'Найдено', 'Отсутствует', 'Процент найденных']
        
        # Форматируем заголовки разделов и таблицу
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value:
                    # Главный заголовок
                    if cell.value == 'Сводка по проверке файлов':
                        cell.font = Font(bold=True, size=14)
                        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                        cell.font = Font(color='FFFFFF', bold=True, size=14)
                    
                    # Заголовки разделов
                    elif cell.value in section_headers[1:]:
                        cell.font = Font(bold=True, size=12)
                        cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
                    
                    # Заголовок таблицы релизов
                    elif isinstance(cell.value, str) and cell.value in table_header:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                        # Применяем форматирование ко всей строке заголовка таблицы
                        for c in range(1, 6):  # 5 колонок в таблице релизов
                            header_cell = ws.cell(row=row_idx, column=c)
                            header_cell.font = Font(bold=True)
                            header_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        # Устанавливаем ширину столбцов
        ws.column_dimensions['A'].width = 35  # Названия релизов или описания
        ws.column_dimensions['B'].width = 15  # Всего файлов
        ws.column_dimensions['C'].width = 15  # Найдено
        ws.column_dimensions['D'].width = 15  # Отсутствует  
        ws.column_dimensions['E'].width = 18  # Процент найденных
    
    # Форматируем лист "Все файлы"
    if 'Все файлы' in writer.sheets and len(all_results_df) > 0:
        ws = writer.sheets['Все файлы']
        
        # Определяем дополнительные цвета для градации
        perfect_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Ярко-зеленый для 100%
        good_green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')     # Светло-зеленый для 90-99%
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')   # Оранжевый для 80-89%
        light_orange = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Светло-оранжевый для 50-79%
        light_red = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')     # Светло-красный для 1-49%
        dark_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')      # Темно-красный для 0%
        
        # Заголовки
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Выделяем строки цветом в зависимости от процента сходства
        for row in range(2, len(all_results_df) + 2):
            try:
                similarity = ws.cell(row=row, column=5).value  # Колонка с процентом сходства
                
                # Определяем цвет в зависимости от процента сходства
                if similarity == 100:
                    # Идеальное совпадение - ярко-зеленый
                    fill_color = perfect_green
                elif similarity >= 90:
                    # Очень хорошее совпадение - светло-зеленый
                    fill_color = good_green
                elif similarity >= 80:
                    # Хорошее совпадение - оранжевый
                    fill_color = orange_fill
                elif similarity >= 50:
                    # Среднее совпадение - светло-оранжевый/золотой
                    fill_color = light_orange
                elif similarity > 0:
                    # Плохое совпадение - светло-красный
                    fill_color = light_red
                else:
                    # Файл не найден - темно-красный
                    fill_color = dark_red
                
                # Применяем цвет ко всей строке
                for col in range(1, 8):  # 7 колонок в таблице
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill_color
                    
                    # Дополнительно выделяем процент сходства жирным шрифтом если < 100%
                    if col == 5 and similarity < 100:  # Колонка с процентом сходства
                        cell.font = Font(bold=True)
                        
            except Exception as e:
                debug_logger.debug(f"⚠️ Ошибка форматирования строки {row}: {e}")
                pass
        
        # Добавляем легенду цветов в конце таблицы
        legend_start_row = len(all_results_df) + 4  # Отступ от основной таблицы
        
        # Заголовок легенды
        legend_header = ws.cell(row=legend_start_row, column=1)
        legend_header.value = "Легенда цветов (по проценту сходства):"
        legend_header.font = Font(bold=True, size=12)
        
        # Элементы легенды
        legend_items = [
            ("100% - Идеальное совпадение", perfect_green),
            ("90-99% - Очень хорошее совпадение", good_green),
            ("80-89% - Хорошее совпадение", orange_fill),
            ("50-79% - Среднее совпадение", light_orange),
            ("1-49% - Плохое совпадение", light_red),
            ("0% - Файл не найден", dark_red)
        ]
        
        for i, (text, color) in enumerate(legend_items):
            row_num = legend_start_row + 1 + i
            cell = ws.cell(row=row_num, column=1)
            cell.value = text
            cell.fill = color
            cell.font = Font(bold=True)
            
            # Объединяем ячейки для лучшего вида легенды
            ws.merge_cells(f'A{row_num}:C{row_num}')
        
        # Автоширина столбцов
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Форматируем лист "Только ошибки" (оригинальная логика)
    if 'Только ошибки' in writer.sheets and len(errors_only_df) > 0:
        ws = writer.sheets['Только ошибки']
        
        # Заголовки
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Проверяем, есть ли это сообщение об отсутствии ошибок
        success_message_row = None
        for row_idx in range(2, len(errors_only_df) + 2):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value == 'Все файлы из Excel найдены в директории':
                success_message_row = row_idx
                break
        
        if success_message_row:
            # Форматируем сообщение об успехе
            success_cell = ws.cell(row=success_message_row, column=1)
            success_cell.font = Font(bold=True, size=12, color='008000')  # Зелёный цвет
            success_cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')  # Светло-зелёный фон
            
            # Объединяем ячейки для сообщения на всю строку
            ws.merge_cells(f'A{success_message_row}:G{success_message_row}')
        else:
            # Выделяем строки цветом в зависимости от процента сходства (только если есть ошибки)
            for row in range(2, len(errors_only_df) + 2):
                try:
                    similarity = ws.cell(row=row, column=5).value  # Колонка с процентом сходства
                    if similarity < 50:
                        for col in range(1, 8):
                            ws.cell(row=row, column=col).fill = red_fill
                    elif similarity < 80:
                        for col in range(1, 8):
                            ws.cell(row=row, column=col).fill = yellow_fill
                except:
                    pass
        
        # Автоширина столбцов
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Форматируем лист "Детальная статистика"
    if 'Детальная статистика' in writer.sheets:
        ws = writer.sheets['Детальная статистика']
        
        # Ищем строку с JWT токеном менеджера и форматируем её
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value == 'JWT токен менеджера':
                    # Форматируем всю строку с JWT токеном менеджера
                    for c in range(1, 4):  # 3 колонки в детальной статистике
                        token_cell = ws.cell(row=row_idx, column=c)
                        token_cell.font = Font(bold=True, size=11)
                        token_cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')  # Светло-оранжевый фон
                    break
        
        # Автоширина столбцов
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Форматируем лист "Рекомендации"
    if 'Рекомендации' in writer.sheets:
        ws = writer.sheets['Рекомендации']
        
        # Автоширина столбцов
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Форматируем лист неиспользованных файлов если есть
    if unused_files_count > 0 and 'Неиспользованные файлы' in writer.sheets:
        ws = writer.sheets['Неиспользованные файлы']
        
        # Заголовки
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Автоширина столбцов
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width


def print_debug_info(filename, normalized_filename):
    """Выводит отладочную информацию о нормализации файла"""
    debug_logger.debug(f"🔍 Искомый файл: '{filename}'")
    debug_logger.debug(f"🔧 Нормализованный файл: '{normalized_filename}'")


def compare_files_interactive():
    """Интерактивная версия для запуска из командной строки"""
    # Получаем абсолютный путь к директории скрипта
    script_dir = Path(__file__).parent.parent
    
    # Пути к файлам
    files_directory = script_dir / 'src' / 'apps' / 'release-parser-5' / 'files'
    excel_file = input("Введите путь к Excel файлу: ")
    
    result = compare_files_with_excel(excel_file, str(files_directory))
    
    if result['success']:
        debug_logger.success(f"✅ {result['message']}")
        debug_logger.info(f"📊 Результаты сохранены в файл: {result['results_file']}")
    else:
        debug_logger.error(f"❌ {result['message']}")


if __name__ == "__main__":
    compare_files_interactive()
